#import openpyxl
#import pandas as pd

#pd.read_xlsx('C:\\users\\Admin\\Desktop\\Play Store Data.xlsx')

#import os

#play_store = open("C:\\users\\Admin\\Desktop\\Play Store Data.xlsx", "r")

#print(play_store.read())

'''
a = open("C:\\users\\Admin\\Desktop\\if.txt","r")

print(a.readline())
a.seek(15)
print(a.readline())

#for x in a:
 #   print(x)
'''


'''
import xlrd

path = "PlayStoreData.xlsx"

Workbook = xlrd.open_workbook(path)
Worksheet = Workbook.sheet_by_index(0)

#print(Worksheet.cell_value(0,0))
#print(Worksheet.ncols)
#print(Worksheet.nrows)
#Worksheet.cell_value(0,0)
#print(Worksheet.row_values(5))
for i in range(Worksheet.ncols):
    print(Worksheet.cell_value(0,i), end=" ")
'''


'''
import pandas as pd

path = pd.read_excel("PlayStoreData.xlsx")

print(path.head(10))
'''


"""
with open("C:\\users\\Admin\\Desktop\\if.txt","r") as file:
    first_line = file.readline()
    for last_line in file:
        pass
print(last_line)
"""


'''
with open("C:\\users\\Admin\\Desktop\\if.txt","r") as file:
    lastline = (list(file)[-1])
print(lastline)
'''

'''
#python2 only
with open("C:\\users\\Admin\\Desktop\\if.txt", "r") as file:
    file.seek(-2,2)
    lastline = file.readline()
print (lastline.strip())
'''


'''
with open("C:\\users\\Admin\\Desktop\\Covid_Sample.txt", "r") as file:
    covidline = (list(file)[8])
print(covidline)
covidline = covidline.split(" ")
#print(covidline)

covidline[0] = covidline[7]
covidline[9] = covidline[13]
covidline[15] = covidline[22]
covidline[25] = covidline[28]
print("Covid_Cases = ",covidline[0])
print("Covid_Death = ",covidline[9])
print("Covid_Processed = ",covidline[15])
print("Covid_Cured = ",covidline[25])
'''

'''
def search_string_in_file(filename,string_to_search):
    lineno = 0
    list_of_result = []
    with open("C:\\users\\Admin\\Desktop\\Covid_Sample.txt", "r") as file:
        for line in file:
            lineno += 1
            if string_to_search in line:
                list_of_result.append((lineno,line.rstrip()))
        return list_of_result

matched = search_string_in_file('sample.txt', 'Covid_Cases')
print("total matched lines : ",len(matched))
for x in matched:
    print('line number = ',x[0],'; line = ',x[1])
'''

'''
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('sheet1')
ws2 = wb.create_sheet('sheet2')
ws3 = wb.create_sheet('sheet3')

wb.save("Myxl.xlsx")
'''









